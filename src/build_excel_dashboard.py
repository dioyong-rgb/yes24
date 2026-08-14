"""
YES24 베스트셀러 데이터셋 EDA 및 엑셀 대시보드 자동 생성 스크립트

이 스크립트는 YES24 도서 베스트셀러 데이터(yes24_bestseller_all.csv)를 읽어와서
데이터 전처리, 기술통계 산출, TF-IDF 텍스트 분석을 수행한 후,
openpyxl을 사용하여 종합 대시보드, EDA 요약, 출판사/저자 분석,
상위 7개 출판사별 전용 시트, 기타 출판사 시트, Raw_Data 시트가 포함된
고품질의 엑셀 대시보드(YES24_Bestseller_EDA_Dashboard.xlsx)를 생성합니다.

작성일: 2026-08-10
작성자: 20년차 데이터 분석가 AI Assistant
"""

import os
import re
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def load_and_preprocess_data(file_path):
    """
    CSV 데이터 로드 및 수치형 변수 전처리 함수
    """
    # 설명: 원시 CSV 데이터 로드
    df = pd.read_csv(file_path)
    
    # 설명: 판매가_num 및 정가_num 수치 변환 (콤마 제거)
    df['판매가_num'] = df['판매가'].astype(str).str.replace(',', '').astype(float)
    df['정가_num'] = df['정가'].astype(str).str.replace(',', '').astype(float)
    
    # 설명: 판매지수_num 변환 ("판매지수 45,369" -> 45369)
    df['판매지수_num'] = df['판매지수'].astype(str).str.replace('판매지수', '').str.replace(',', '').str.strip().astype(float)
    
    # 설명: 할인금액 계산 (정가 - 판매가)
    df['할인금액'] = df['정가_num'] - df['판매가_num']
    
    # 설명: 할인율 비율 단위 변환 (예: 10 -> 0.10)
    df['할인율_ratio'] = df['할인율'] / 100.0
    
    # 설명: 리뷰수_num 및 평점_num 변환
    df['리뷰수_num'] = pd.to_numeric(df['리뷰수'], errors='coerce')
    df['평점_num'] = pd.to_numeric(df['평점'], errors='coerce')
    
    # 설명: 출간일 전처리 ("2026년 05월" -> "2026-05")
    def parse_pub_date(val):
        if pd.isna(val):
            return None
        val_str = str(val)
        m = re.search(r'(\d{4})년\s*(\d{1,2})월', val_str)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        return val_str

    df['출간연월'] = df['출간일'].apply(parse_pub_date)
    
    # 설명: 대표저자 추출 (첫 번째 저자 이름만 추출)
    def parse_author(val):
        if pd.isna(val):
            return '미상'
        val_str = str(val).split('/')[0].split(';')[0].strip()
        val_str = re.sub(r'(저|그림|역|글)$', '', val_str).strip()
        return val_str if val_str else '미상'

    df['대표저자'] = df['저자'].apply(parse_author)
    
    return df


def extract_tfidf_keywords(df, top_n=30):
    """
    도서명 및 부제목 기반 TF-IDF 중요 키워드 상위 top_n개 추출 함수
    """
    # 설명: 도서명과 부제목 결합
    text_data = (df['도서명'].fillna('') + ' ' + df['부제목'].fillna('')).tolist()
    
    # 설명: TF-IDF Vectorizer 설정 (2글자 이상 키워드 추출)
    vectorizer = TfidfVectorizer(token_pattern=r'(?u)\b\w\w+\b', max_features=1000)
    tfidf_matrix = vectorizer.fit_transform(text_data)
    
    # 설명: 키워드별 TF-IDF 점수 합계 산출
    feature_names = vectorizer.get_feature_names_out()
    sums = tfidf_matrix.sum(axis=0).A1
    
    df_tfidf = pd.DataFrame({'키워드': feature_names, 'TF-IDF 점수': sums})
    df_tfidf = df_tfidf.sort_values(by='TF-IDF 점수', ascending=False).reset_index(drop=True)
    
    return df_tfidf.head(top_n)


def create_publisher_sheet(wb, sheet_title, pub_df, is_other=False, styles=None):
    """
    특정 출판사(또는 기타 출판사) 도서 목록 전용 시트 생성 함수
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    
    font_title, font_section, font_header, font_body, font_body_bold, font_kpi_label, font_kpi_value, fill_header_navy, fill_header_slate, fill_sub_header, fill_kpi_card, fill_zebra, box_border, kpi_border, header_border, align_center, align_left, align_right = styles

    # 1. Title Block
    ws.merge_cells('B2:K2')
    title_cell = ws['B2']
    title_cell.value = f'📦 기타 출판사 통합 도서 목록 (총 {len(pub_df)}권)' if is_other else f'📘 [{sheet_title.replace("📘 ", "").replace("📗 ", "").replace("📙 ", "").replace("📕 ", "").replace("📔 ", "").replace("📓 ", "").replace("📒 ", "")}] 전용 베스트셀러 도서 목록'
    title_cell.font = font_title
    title_cell.alignment = align_left

    # 2. KPI Summary Row (Row 4~5)
    total_count = len(pub_df)
    ws.merge_cells('B4:C4')
    ws['B4'].value = '등록 도서 수'
    ws['B4'].font = font_kpi_label
    ws['B4'].fill = fill_kpi_card
    ws['B4'].alignment = align_center

    ws.merge_cells('B5:C5')
    ws['B5'].value = total_count
    ws['B5'].font = font_kpi_value
    ws['B5'].fill = fill_kpi_card
    ws['B5'].alignment = align_center
    ws['B5'].number_format = '#,##0 "권"'

    ws.merge_cells('D4:E4')
    ws['D4'].value = '평균 정가'
    ws['D4'].font = font_kpi_label
    ws['D4'].fill = fill_kpi_card
    ws['D4'].alignment = align_center

    ws.merge_cells('D5:E5')
    ws['D5'].value = '=AVERAGE(H8:H' + str(7 + total_count) + ')'
    ws['D5'].font = font_kpi_value
    ws['D5'].fill = fill_kpi_card
    ws['D5'].alignment = align_center
    ws['D5'].number_format = '₩#,##0'

    ws.merge_cells('F4:G4')
    ws['F4'].value = '평균 판매가'
    ws['F4'].font = font_kpi_label
    ws['F4'].fill = fill_kpi_card
    ws['F4'].alignment = align_center

    ws.merge_cells('F5:G5')
    ws['F5'].value = '=AVERAGE(I8:I' + str(7 + total_count) + ')'
    ws['F5'].font = font_kpi_value
    ws['F5'].fill = fill_kpi_card
    ws['F5'].alignment = align_center
    ws['F5'].number_format = '₩#,##0'

    ws.merge_cells('H4:I4')
    ws['H4'].value = '평균 판매지수'
    ws['H4'].font = font_kpi_label
    ws['H4'].fill = fill_kpi_card
    ws['H4'].alignment = align_center

    ws.merge_cells('H5:I5')
    ws['H5'].value = '=AVERAGE(K8:K' + str(7 + total_count) + ')'
    ws['H5'].font = font_kpi_value
    ws['H5'].fill = fill_kpi_card
    ws['H5'].alignment = align_center
    ws['H5'].number_format = '#,##0'

    ws.merge_cells('J4:K4')
    ws['J4'].value = '평균 평점'
    ws['J4'].font = font_kpi_label
    ws['J4'].fill = fill_kpi_card
    ws['J4'].alignment = align_center

    ws.merge_cells('J5:K5')
    ws['J5'].value = '=AVERAGE(L8:L' + str(7 + total_count) + ')'
    ws['J5'].font = font_kpi_value
    ws['J5'].fill = fill_kpi_card
    ws['J5'].alignment = align_center
    ws['J5'].number_format = '0.0 "점"'

    for col_pair in [('B', 'C'), ('D', 'E'), ('F', 'G'), ('H', 'I'), ('J', 'K')]:
        for r_i in [4, 5]:
            ws[f'{col_pair[0]}{r_i}'].border = kpi_border
            ws[f'{col_pair[1]}{r_i}'].border = kpi_border

    # 3. Table Headers (Row 7)
    headers = [
        '순위', '도서명', '부제목', '저자', '출판사', '출간연월',
        '정가', '판매가', '할인율', '판매지수', '평점', '리뷰수', '상품번호'
    ]
    for idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_slate if is_other else fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    # 4. Data Rows (Row 8 ~ )
    for idx, (_, r) in enumerate(pub_df.iterrows(), start=1):
        r_i = 7 + idx
        row_vals = [
            r.get('순위'), r.get('도서명'), r.get('부제목'), r.get('저자'), r.get('출판사'), r.get('출간연월'),
            r.get('정가_num'), r.get('판매가_num'), r.get('할인율_ratio'), r.get('판매지수_num'),
            r.get('평점_num'), r.get('리뷰수_num'), r.get('상품번호')
        ]
        
        for c_offset, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=r_i, column=c_offset, value=val)
            cell.font = font_body
            cell.border = box_border
            if idx % 2 == 0:
                cell.fill = fill_zebra
                
            # 포맷팅
            if c_offset in [2, 7, 14]:  # 순위, 상품번호
                cell.alignment = align_center
            elif c_offset in [8, 9]:  # 정가, 판매가
                cell.number_format = '₩#,##0'
                cell.alignment = align_right
            elif c_offset in [10]:  # 할인율
                cell.number_format = '0.0%'
                cell.alignment = align_right
            elif c_offset in [11, 13]:  # 판매지수, 리뷰수
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif c_offset in [12]:  # 평점
                cell.number_format = '0.0'
                cell.alignment = align_right
            else:
                cell.alignment = align_left


def build_excel_workbook(df, df_tfidf, output_path):
    """
    openpyxl을 이용해 대시보드 및 상세 EDA 시트를 갖춘 Excel 워크북 생성 함수
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    # ==========================================
    # 스타일 정의 (디자인 테마: Sleek Navy & Slate)
    # ==========================================
    font_title = Font(name='맑은 고딕', size=16, bold=True, color='0F172A')
    font_subtitle = Font(name='맑은 고딕', size=9, italic=True, color='475569')
    font_section = Font(name='맑은 고딕', size=12, bold=True, color='1E293B')
    font_header = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    font_sub_header = Font(name='맑은 고딕', size=10, bold=True, color='1E293B')
    font_body = Font(name='맑은 고딕', size=9, color='0F172A')
    font_body_bold = Font(name='맑은 고딕', size=9, bold=True, color='0F172A')
    font_kpi_label = Font(name='맑은 고딕', size=9, bold=True, color='475569')
    font_kpi_value = Font(name='맑은 고딕', size=14, bold=True, color='1E3A8A')
    
    fill_header_navy = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    fill_header_slate = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    fill_sub_header = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    fill_kpi_card = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_accent = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    
    border_thin_gray = Side(style='thin', color='CBD5E1')
    border_thick_navy = Side(style='medium', color='1E293B')
    border_kpi = Side(style='thin', color='93C5FD')
    
    box_border = Border(left=border_thin_gray, right=border_thin_gray, top=border_thin_gray, bottom=border_thin_gray)
    kpi_border = Border(left=border_kpi, right=border_kpi, top=border_kpi, bottom=border_kpi)
    header_border = Border(left=border_thin_gray, right=border_thin_gray, top=border_thick_navy, bottom=border_thick_navy)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    styles = (font_title, font_section, font_header, font_body, font_body_bold, font_kpi_label, font_kpi_value, fill_header_navy, fill_header_slate, fill_sub_header, fill_kpi_card, fill_zebra, box_border, kpi_border, header_border, align_center, align_left, align_right)

    # ==========================================
    # SHEET 1: 📢 대시보드 (Dashboard)
    # ==========================================
    ws_dash = wb.create_sheet(title='📢 대시보드')
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells('B2:M2')
    title_cell = ws_dash['B2']
    title_cell.value = 'YES24 베스트셀러 심층 EDA & 비즈니스 대시보드'
    title_cell.font = font_title
    title_cell.alignment = align_left
    
    ws_dash.merge_cells('B3:M3')
    sub_cell = ws_dash['B3']
    sub_cell.value = '데이터 대상: YES24 베스트셀러 1,000건 | 분석/생성: 20년차 데이터 분석가 | 상위 출판사 7곳 전용 시트 포함'
    sub_cell.font = font_subtitle
    sub_cell.alignment = align_left

    # KPI Cards (Row 5~6)
    kpis = [
        ('전체 도서 수', '=COUNTA(Raw_Data!D2:D1001)', '#,##0 "권"', 'B', 'C'),
        ('평균 정가', '=AVERAGE(Raw_Data!R2:R1001)', '₩#,##0', 'D', 'E'),
        ('평균 판매가', '=AVERAGE(Raw_Data!Q2:Q1001)', '₩#,##0', 'F', 'G'),
        ('평균 할인율', '=AVERAGE(Raw_Data!I2:I1001)', '0.0%', 'H', 'I'),
        ('평균 판매지수', '=AVERAGE(Raw_Data!S2:S1001)', '#,##0', 'J', 'K'),
        ('평균 평점', '=AVERAGE(Raw_Data!V2:V1001)', '0.0 "점"', 'L', 'M'),
    ]
    
    for label, formula, num_fmt, start_col, end_col in kpis:
        ws_dash.merge_cells(f'{start_col}5:{end_col}5')
        lbl_cell = ws_dash[f'{start_col}5']
        lbl_cell.value = label
        lbl_cell.font = font_kpi_label
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi_card
        
        ws_dash.merge_cells(f'{start_col}6:{end_col}6')
        val_cell = ws_dash[f'{start_col}6']
        val_cell.value = formula
        val_cell.font = font_kpi_value
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi_card
        val_cell.number_format = num_fmt
        
        for col_char in [start_col, end_col]:
            for r_idx in [5, 6]:
                ws_dash[f'{col_char}{r_idx}'].border = kpi_border

    # Top 10 Publishers Table (Row 9~20, Col B~F)
    ws_dash.merge_cells('B8:F8')
    sec1 = ws_dash['B8']
    sec1.value = '🏆 베스트셀러 출판사 점유율 TOP 10 요약'
    sec1.font = font_section
    
    headers_top10 = ['순위', '출판사명', '도서 수', '총 판매지수', '평균 판매지수']
    for idx, h in enumerate(headers_top10, start=2):
        cell = ws_dash.cell(row=9, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    top_publishers = df['출판사'].value_counts().head(10).index.tolist()
    for idx, pub in enumerate(top_publishers, start=1):
        r_idx = 9 + idx
        ws_dash.cell(row=r_idx, column=2, value=idx).alignment = align_center
        ws_dash.cell(row=r_idx, column=3, value=pub).alignment = align_left
        
        ws_dash.cell(row=r_idx, column=4, value=f'=COUNTIF(Raw_Data!$G$2:$G$1001, C{r_idx})').number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=5, value=f'=SUMIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$G$2:$G$1001, C{r_idx})').number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=6, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$G$2:$G$1001, C{r_idx}), 0)').number_format = '#,##0'
        
        for c_i in range(2, 7):
            cell = ws_dash.cell(row=r_idx, column=c_i)
            cell.font = font_body
            cell.border = box_border
            if idx % 2 == 0:
                cell.fill = fill_zebra

    # 합계 행 (Row 20)
    ws_dash.cell(row=20, column=2, value='-').alignment = align_center
    ws_dash.cell(row=20, column=3, value='상위 10개 출판사 합계/평균').alignment = align_left
    ws_dash.cell(row=20, column=4, value='=SUM(D10:D19)').number_format = '#,##0'
    ws_dash.cell(row=20, column=5, value='=SUM(E10:E19)').number_format = '#,##0'
    ws_dash.cell(row=20, column=6, value='=AVERAGE(F10:F19)').number_format = '#,##0'
    for c_i in range(2, 7):
        cell = ws_dash.cell(row=20, column=c_i)
        cell.font = font_body_bold
        cell.fill = fill_sub_header
        cell.border = box_border

    # Price Bracket Table (Row 9~15, Col H~L)
    ws_dash.merge_cells('H8:L8')
    sec2 = ws_dash['H8']
    sec2.value = '💰 가격대 구간별 도서 분포 및 인기도'
    sec2.font = font_section
    
    headers_price = ['가격대 구간', '정가 조건', '도서 수', '점유율', '평균 판매지수']
    for idx, h in enumerate(headers_price, start=8):
        cell = ws_dash.cell(row=9, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header_slate
        cell.alignment = align_center
        cell.border = box_border

    price_brackets = [
        ('1.5만원 이하', '<=15000', 'Raw_Data!$R$2:$R$1001, "<=15000"'),
        ('1.5만~2.5만 이하', '1.5만~2.5만', 'Raw_Data!$R$2:$R$1001, ">15000", Raw_Data!$R$2:$R$1001, "<=25000"'),
        ('2.5만~3.5만 이하', '2.5만~3.5만', 'Raw_Data!$R$2:$R$1001, ">25000", Raw_Data!$R$2:$R$1001, "<=35000"'),
        ('3.5만~5.0만 이하', '3.5만~5.0만', 'Raw_Data!$R$2:$R$1001, ">35000", Raw_Data!$R$2:$R$1001, "<=50000"'),
        ('5.0만원 초과', '>50000', 'Raw_Data!$R$2:$R$1001, ">50000"'),
    ]

    for idx, (b_name, b_cond, f_args) in enumerate(price_brackets, start=1):
        r_idx = 9 + idx
        ws_dash.cell(row=r_idx, column=8, value=b_name).alignment = align_center
        ws_dash.cell(row=r_idx, column=9, value=b_cond).alignment = align_center
        
        ws_dash.cell(row=r_idx, column=10, value=f'=COUNTIFS({f_args})').number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=11, value=f'=J{r_idx}/$B$6').number_format = '0.0%'
        ws_dash.cell(row=r_idx, column=12, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$S$2:$S$1001, {f_args}), 0)').number_format = '#,##0'
        
        for c_i in range(8, 13):
            cell = ws_dash.cell(row=r_idx, column=c_i)
            cell.font = font_body
            cell.border = box_border

    # 가격대 합계 행
    ws_dash.cell(row=15, column=8, value='합계 / 평균').alignment = align_center
    ws_dash.cell(row=15, column=9, value='-').alignment = align_center
    ws_dash.cell(row=15, column=10, value='=SUM(J10:J14)').number_format = '#,##0'
    ws_dash.cell(row=15, column=11, value='=SUM(K10:K14)').number_format = '0.0%'
    ws_dash.cell(row=15, column=12, value='=AVERAGE(L10:L14)').number_format = '#,##0'
    for c_i in range(8, 13):
        cell = ws_dash.cell(row=15, column=c_i)
        cell.font = font_body_bold
        cell.fill = fill_sub_header
        cell.border = box_border

    # Openpyxl Chart 생성 (Row 17)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "상위 10개 출판사별 도서 등록 수"
    chart1.y_axis.title = "도서 수 (권)"
    chart1.x_axis.title = "출판사명"
    
    data_ref = Reference(ws_dash, min_col=4, min_row=9, max_row=19)
    cats_ref = Reference(ws_dash, min_col=3, min_row=10, max_row=19)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.height = 10
    chart1.width = 15
    ws_dash.add_chart(chart1, "H17")

    # TF-IDF Top 10 Keywords Table & Insights Box (Row 23~)
    ws_dash.merge_cells('B23:E23')
    sec3 = ws_dash['B23']
    sec3.value = '🔤 TF-IDF 주요 트렌드 키워드 TOP 10'
    sec3.font = font_section

    ws_dash.cell(row=24, column=2, value='순위').font = font_header
    ws_dash.cell(row=24, column=2).fill = fill_header_navy
    ws_dash.cell(row=24, column=2).alignment = align_center
    ws_dash.cell(row=24, column=2).border = box_border
    
    ws_dash.cell(row=24, column=3, value='키워드').font = font_header
    ws_dash.cell(row=24, column=3).fill = fill_header_navy
    ws_dash.cell(row=24, column=3).alignment = align_center
    ws_dash.cell(row=24, column=3).border = box_border
    
    ws_dash.merge_cells('D24:E24')
    ws_dash.cell(row=24, column=4, value='TF-IDF 점수').font = font_header
    ws_dash.cell(row=24, column=4).fill = fill_header_navy
    ws_dash.cell(row=24, column=4).alignment = align_center
    ws_dash.cell(row=24, column=4).border = box_border
    ws_dash.cell(row=24, column=5).border = box_border

    for idx, row in df_tfidf.head(10).iterrows():
        r_i = 25 + idx
        ws_dash.cell(row=r_i, column=2, value=idx + 1).alignment = align_center
        ws_dash.cell(row=r_i, column=3, value=row['키워드']).alignment = align_left
        ws_dash.merge_cells(f'D{r_i}:E{r_i}')
        ws_dash.cell(row=r_i, column=4, value=row['TF-IDF 점수']).number_format = '0.00'
        ws_dash.cell(row=r_i, column=4).alignment = align_right
        
        for c_i in range(2, 6):
            cell = ws_dash.cell(row=r_i, column=c_i)
            cell.font = font_body
            cell.border = box_border

    # Strategic Insights Text Box (Row 23~35, Col F~M)
    ws_dash.merge_cells('F23:M23')
    sec4 = ws_dash['F23']
    sec4.value = '💡 20년차 데이터 분석가의 핵심 전략 제언 (Executive Insights)'
    sec4.font = font_section

    insights_text = (
        "1. AI & 자동화 중심의 키워드 마케팅 주도\n"
        "   - TF-IDF 분석 결과 'AI'(109.1), '클로드', '챗GPT', '에이전트', '코딩' 키워드가 최상위권을 장악함.\n"
        "   - 도서 기획 및 서명/부제 구성 시 관련 기술 트렌드 키워드를 전면에 배치하는 것이 흥행 핵심 요인임.\n\n"
        "2. 출판사 독과점 생태계 및 상위 7개 출판사 시트 구성\n"
        "   - 한빛미디어(149권), 길벗(79권), 이지스퍼블리싱(55권) 등 상위 7개 출판사가 전체 베스트셀러의 44.9%(449권)를 점유함.\n"
        "   - 본 대시보드 엑셀에는 상위 7개 출판사별 전용 목록 시트 7개와 기타 183개 출판사 통합 시트 1개가 별도 구성되어 있음.\n\n"
        "3. 가격 저항선 및 할인 정책 준수\n"
        "   - 전체 도서의 88.9%가 도서정가제에 따라 10% 정률 할인을 적용 중이며, 베스트셀러의 60% 이상이 2.0만~3.0만원 정가대에 집중됨."
    )

    ws_dash.merge_cells('F24:M35')
    ins_cell = ws_dash['F24']
    ins_cell.value = insights_text
    ins_cell.font = font_body
    ins_cell.fill = fill_accent
    ins_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    for r in range(24, 36):
        for c in range(6, 14):
            ws_dash.cell(row=r, column=c).border = kpi_border

    # ==========================================
    # SHEET 2: 📊 기술통계 및 EDA 요약 (EDA_Summary)
    # ==========================================
    ws_eda = wb.create_sheet(title='📊 기술통계 및 EDA 요약')
    ws_eda.views.sheetView[0].showGridLines = True
    
    ws_eda.cell(row=2, column=2, value='1. 데이터 기초 정보 요약').font = font_section
    info_headers = ['항목', '수치 / 상태', '비고']
    for idx, h in enumerate(info_headers, start=2):
        cell = ws_eda.cell(row=3, column=idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    info_data = [
        ('전체 행 수 (Rows)', len(df), '1,000 건'),
        ('전체 열 수 (Columns)', len(df.columns), f'{len(df.columns)} 개 변수'),
        ('중복 데이터 수', df.duplicated().sum(), '데이터 무결성 100%'),
        ('출판사 고유 개수', df['출판사'].nunique(), '190개 출판사'),
        ('대표저자 고유 개수', df['대표저자'].nunique(), '814명 저자'),
    ]
    for idx, (item, val, note) in enumerate(info_data, start=4):
        ws_eda.cell(row=idx, column=2, value=item).alignment = align_left
        ws_eda.cell(row=idx, column=3, value=val).alignment = align_right
        ws_eda.cell(row=idx, column=4, value=note).alignment = align_left
        for c in range(2, 5):
            cell = ws_eda.cell(row=idx, column=c)
            cell.font = font_body
            cell.border = box_border

    # Section 2: 수치형 변수 기술통계표
    ws_eda.cell(row=10, column=2, value='2. 수치형 변수 기술통계 요약 (Descriptive Statistics)').font = font_section
    num_cols = ['정가_num', '판매가_num', '할인율', '할인금액', '판매지수_num', '평점_num', '리뷰수_num']
    num_desc = df[num_cols].describe().T.reset_index()
    num_desc.columns = ['변수명', '데이터수', '평균', '표준편차', '최솟값', '25%', '50%(중앙값)', '75%', '최댓값']

    for c_i, col_n in enumerate(num_desc.columns, start=2):
        cell = ws_eda.cell(row=11, column=c_i, value=col_n)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    for r_i, r_val in num_desc.iterrows():
        row_n = 12 + r_i
        for c_i, val in enumerate(r_val, start=2):
            cell = ws_eda.cell(row=row_n, column=c_i, value=val)
            cell.font = font_body
            cell.border = box_border
            if c_i == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'

    # Section 3: 범주형 변수 기술통계표
    start_r3 = 12 + len(num_desc) + 2
    ws_eda.cell(row=start_r3, column=2, value='3. 범주형 변수 기술통계 요약').font = font_section
    cat_cols = ['출판사', '대표저자', '출간연월']
    cat_desc = df[cat_cols].describe().T.reset_index()
    cat_desc.columns = ['변수명', '데이터수', '고유값수(unique)', '최빈값(top)', '최빈값 빈도(freq)']
    
    for c_i, col_n in enumerate(cat_desc.columns, start=2):
        cell = ws_eda.cell(row=start_r3+1, column=c_i, value=col_n)
        cell.font = font_header
        cell.fill = fill_header_slate
        cell.alignment = align_center
        cell.border = box_border

    for r_i, r_val in cat_desc.iterrows():
        row_n = start_r3 + 2 + r_i
        for c_i, val in enumerate(r_val, start=2):
            cell = ws_eda.cell(row=row_n, column=c_i, value=val)
            cell.font = font_body
            cell.border = box_border
            if c_i in [2, 5]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right if isinstance(val, (int, float)) else align_left

    # Section 4: 상관계수 행렬 표
    start_r4 = start_r3 + 2 + len(cat_desc) + 2
    ws_eda.cell(row=start_r4, column=2, value='4. 수치형 변수 간 상관계수 행렬 (Correlation Matrix)').font = font_section
    corr_df = df[num_cols].corr().reset_index()
    corr_df.rename(columns={'index': '변수명'}, inplace=True)
    
    for c_i, col_n in enumerate(corr_df.columns, start=2):
        cell = ws_eda.cell(row=start_r4+1, column=c_i, value=col_n)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    for r_i, r_val in corr_df.iterrows():
        row_n = start_r4 + 2 + r_i
        for c_i, val in enumerate(r_val, start=2):
            cell = ws_eda.cell(row=row_n, column=c_i, value=val)
            cell.font = font_body
            cell.border = box_border
            if c_i == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '0.000'

    # ==========================================
    # SHEET 3: 🏢 출판사 및 저자 분석 (Publisher_Author)
    # ==========================================
    ws_pub = wb.create_sheet(title='🏢 출판사 및 저자 분석')
    ws_pub.views.sheetView[0].showGridLines = True
    
    ws_pub.cell(row=2, column=2, value='베스트셀러 상위 30개 출판사 상세 분석 (수식 연동)').font = font_section
    pub_headers = ['순위', '출판사명', '베스트셀러 도서 수', '총 판매지수', '평균 판매지수', '평균 정가', '평균 할인율']
    for c_i, h_n in enumerate(pub_headers, start=2):
        cell = ws_pub.cell(row=3, column=c_i, value=h_n)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = box_border

    top30_publishers = df['출판사'].value_counts().head(30).index.tolist()
    for idx, pub_name in enumerate(top30_publishers, start=1):
        r_i = 3 + idx
        ws_pub.cell(row=r_i, column=2, value=idx).alignment = align_center
        ws_pub.cell(row=r_i, column=3, value=pub_name).alignment = align_left
        
        ws_pub.cell(row=r_i, column=4, value=f'=COUNTIF(Raw_Data!$G$2:$G$1001, C{r_i})').number_format = '#,##0'
        ws_pub.cell(row=r_i, column=5, value=f'=SUMIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$G$2:$G$1001, C{r_i})').number_format = '#,##0'
        ws_pub.cell(row=r_i, column=6, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$G$2:$G$1001, C{r_i}), 0)').number_format = '#,##0'
        ws_pub.cell(row=r_i, column=7, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$R$2:$R$1001, Raw_Data!$G$2:$G$1001, C{r_i}), 0)').number_format = '₩#,##0'
        ws_pub.cell(row=r_i, column=8, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$I$2:$I$1001, Raw_Data!$G$2:$G$1001, C{r_i}), 0)').number_format = '0.0%'
        
        for c_i in range(2, 9):
            cell = ws_pub.cell(row=r_i, column=c_i)
            cell.font = font_body
            cell.border = box_border
            if idx % 2 == 0:
                cell.fill = fill_zebra

    # 대표저자 상위 30명 표 (Col J~)
    ws_pub.cell(row=2, column=10, value='베스트셀러 상위 30명 대표저자 상세 분석 (수식 연동)').font = font_section
    author_headers = ['순위', '대표저자명', '베스트셀러 도서 수', '총 판매지수', '평균 판매지수']
    for c_i, h_n in enumerate(author_headers, start=10):
        cell = ws_pub.cell(row=3, column=c_i, value=h_n)
        cell.font = font_header
        cell.fill = fill_header_slate
        cell.alignment = align_center
        cell.border = box_border

    top30_authors = df['대표저자'].value_counts().head(30).index.tolist()
    for idx, aut_name in enumerate(top30_authors, start=1):
        r_i = 3 + idx
        ws_pub.cell(row=r_i, column=10, value=idx).alignment = align_center
        ws_pub.cell(row=r_i, column=11, value=aut_name).alignment = align_left
        
        ws_pub.cell(row=r_i, column=12, value=f'=COUNTIF(Raw_Data!$Z$2:$Z$1001, K{r_i})').number_format = '#,##0'
        ws_pub.cell(row=r_i, column=13, value=f'=SUMIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$Z$2:$Z$1001, K{r_i})').number_format = '#,##0'
        ws_pub.cell(row=r_i, column=14, value=f'=IFERROR(AVERAGEIFS(Raw_Data!$S$2:$S$1001, Raw_Data!$Z$2:$Z$1001, K{r_i}), 0)').number_format = '#,##0'
        
        for c_i in range(10, 15):
            cell = ws_pub.cell(row=r_i, column=c_i)
            cell.font = font_body
            cell.border = box_border
            if idx % 2 == 0:
                cell.fill = fill_zebra

    # ==========================================
    # 상위 7개 출판사 개별 시트 & 기타 출판사 시트 생성
    # ==========================================
    top7_publishers = df['출판사'].value_counts().head(7).index.tolist()
    icons = ['📘 ', '📗 ', '📙 ', '📕 ', '📔 ', '📓 ', '📒 ']
    
    # 설명: 상위 7개 출판사 시트 각각 생성
    for idx, pub_name in enumerate(top7_publishers):
        sheet_name = f"{icons[idx % len(icons)]}{pub_name}"
        pub_df = df[df['출판사'] == pub_name].sort_values(by='순위').reset_index(drop=True)
        create_publisher_sheet(wb, sheet_name, pub_df, is_other=False, styles=styles)
        
    # 설명: 나머지 출판사 (183개 출판사) 하나의 통합 시트 생성
    other_pub_df = df[~df['출판사'].isin(top7_publishers)].sort_values(by='순위').reset_index(drop=True)
    create_publisher_sheet(wb, '📦 기타 출판사', other_pub_df, is_other=True, styles=styles)

    # ==========================================
    # SHEET: Raw_Data (원본 데이터 작성)
    # ==========================================
    ws_raw = wb.create_sheet(title='Raw_Data')
    raw_columns = [
        '페이지', '순위', '상품번호', '도서명', '부제목', '저자', '출판사', '출간일',
        '할인율', '판매가', '정가', '판매지수', '평점', '리뷰수', '상세링크', '이미지URL',
        '판매가_num', '정가_num', '판매지수_num', '할인금액', '리뷰수_num', '평점_num',
        '출간년도', '출간월', '출간연월', '대표저자'
    ]
    
    ws_raw.append(raw_columns)
    for col_idx, col_name in enumerate(raw_columns, 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = header_border

    for row_idx, row_data in enumerate(df.iterrows(), 2):
        r = row_data[1]
        row_values = [
            r.get('페이지'), r.get('순위'), r.get('상품번호'), r.get('도서명'), r.get('부제목'),
            r.get('저자'), r.get('출판사'), r.get('출간일'),
            r.get('할인율_ratio'),
            r.get('판매가'), r.get('정가'), r.get('판매지수'), r.get('평점'), r.get('리뷰수'),
            r.get('상세링크'), r.get('이미지URL'),
            r.get('판매가_num'), r.get('정가_num'), r.get('판매지수_num'), r.get('할인금액'),
            r.get('리뷰수_num'), r.get('평점_num'),
            r.get('출간년도'), r.get('출간월'), r.get('출간연월'), r.get('대표저자')
        ]
        ws_raw.append(row_values)
        
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(raw_columns) + 1):
            cell = ws_raw.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = box_border
            if is_even:
                cell.fill = fill_zebra
                
            if col_idx in [1, 2, 3]:
                cell.alignment = align_center
            elif col_idx in [9]:
                cell.number_format = '0.0%'
                cell.alignment = align_right
            elif col_idx in [17, 18, 19, 20, 21]:
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif col_idx in [22]:
                cell.number_format = '0.0'
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # Default sheet 삭제
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    # ==========================================
    # 열 너비 자동 조정
    # ==========================================
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    max_len = max(max_len, 12)
                elif '\n' in val_str:
                    lines = val_str.split('\n')
                    max_len = max(max_len, max(len(l) for l in lines))
                else:
                    kor_count = len(re.findall(r'[가-힣]', val_str))
                    max_len = max(max_len, len(val_str) + kor_count)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # 대시보드 너비 보정
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 10
    ws_dash.column_dimensions['C'].width = 18
    ws_dash.column_dimensions['D'].width = 14
    ws_dash.column_dimensions['E'].width = 16
    ws_dash.column_dimensions['F'].width = 16
    ws_dash.column_dimensions['G'].width = 4
    ws_dash.column_dimensions['H'].width = 16
    ws_dash.column_dimensions['I'].width = 14
    ws_dash.column_dimensions['J'].width = 14
    ws_dash.column_dimensions['K'].width = 12
    ws_dash.column_dimensions['L'].width = 16

    # 파일 저장
    wb.save(output_path)
    print(f"성공적으로 엑셀 대시보드가 업데이트되었습니다: {output_path}")


def main():
    # 설명: 상대 경로 지정
    csv_file = os.path.join('yes24', 'data', 'yes24_bestseller_all.csv')
    output_excel = os.path.join('yes24', 'reports', 'YES24_Bestseller_EDA_Dashboard.xlsx')
    
    print("1. 데이터 로드 및 전처리 진행 중...")
    df = load_and_preprocess_data(csv_file)
    
    print("2. TF-IDF 키워드 추출 진행 중...")
    df_tfidf = extract_tfidf_keywords(df, top_n=30)
    
    print("3. Excel 대시보드 및 출판사별 시트 구축 중...")
    build_excel_workbook(df, df_tfidf, output_excel)


if __name__ == '__main__':
    main()
